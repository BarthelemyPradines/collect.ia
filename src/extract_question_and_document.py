import re
from pathlib import Path

import openpyxl
import pandas as pd


def _parse_documents(cell_value: str) -> list[str]:
    """Parse a cell like '1. doc1 \n2. doc2' into ['doc1', 'doc2']."""
    parts = re.split(r"\d+\.?\s*", cell_value)
    return [p.strip() for p in parts if p.strip()]


def _build_category_map(ws: openpyxl.worksheet.worksheet.Worksheet, category_row: int) -> dict[int, str]:
    """Map each column index to its category based on merged cells in the category row."""
    col_to_category: dict[int, str] = {}
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= category_row <= merged_range.max_row:
            value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            if value and str(value).strip():
                label = str(value).strip()
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    col_to_category[c] = label

    # Also pick up non-merged cells that have a value on the category row.
    for col_idx in range(1, ws.max_column + 1):
        if col_idx not in col_to_category:
            cell_value = ws.cell(row=category_row, column=col_idx).value
            if cell_value and str(cell_value).strip():
                col_to_category[col_idx] = str(cell_value).strip()

    return col_to_category


def extract_questions_and_documents(
    file_path: Path,
    sheet_name: str,
    start_col: int,
    document_row: int,
    question_row: int,
    category_row: int | None = None,
    structured_output_row: int | None = None,
) -> pd.DataFrame:
    """Read an Excel sheet and return a DataFrame pairing each question with its documents.

    Args:
        file_path: Path to the .xlsx file.
        sheet_name: Name of the sheet to read.
        start_col: First column (1-indexed) containing data.
        document_row: Row number (1-indexed) containing documents.
        question_row: Row number (1-indexed) containing questions.
        category_row: Optional row number (1-indexed) between documents and questions
                      that may contain merged cells with category labels.
        structured_output_row: Optional row number (1-indexed) containing structured
                               output specs (e.g. "Bool", "str=-COMPLET-INCOMPLET-ABSENT-SO").

    Returns:
        DataFrame with columns: question, category (if applicable), structured_output,
        question_ref, document_ref, 1, 2, ...
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    col_to_category: dict[int, str] = {}
    if category_row is not None:
        col_to_category = _build_category_map(ws, category_row)

    rows: list[dict] = []
    max_docs = 0

    for col_idx in range(start_col, ws.max_column + 1):
        question_cell = ws.cell(row=question_row, column=col_idx)
        document_cell = ws.cell(row=document_row, column=col_idx)

        question = question_cell.value
        document_raw = document_cell.value

        if question is None or document_raw is None:
            continue

        question = str(question).strip()
        document_raw = str(document_raw).strip()

        if not question or not document_raw:
            continue

        docs = _parse_documents(document_raw)
        if not docs:
            continue

        col_letter = openpyxl.utils.get_column_letter(col_idx)
        row_data: dict = {
            "question": question,
            "question_ref": f"{col_letter}{question_row}",
            "document_ref": f"{col_letter}{document_row}",
        }

        if category_row is not None:
            row_data["category"] = col_to_category.get(col_idx)

        if structured_output_row is not None:
            spec = ws.cell(row=structured_output_row, column=col_idx).value
            row_data["structured_output"] = str(spec).strip() if spec else None

        for i, doc in enumerate(docs, start=1):
            row_data[str(i)] = doc

        max_docs = max(max_docs, len(docs))
        rows.append(row_data)

    wb.close()

    if not rows:
        raise ValueError("No valid question/document pairs found in the sheet.")

    doc_columns = [str(i) for i in range(1, max_docs + 1)]
    base_columns = ["question"]
    if category_row is not None:
        base_columns.append("category")
    if structured_output_row is not None:
        base_columns.append("structured_output")
    base_columns += ["question_ref", "document_ref"]

    df = pd.DataFrame(rows, columns=base_columns + doc_columns)
    return df
